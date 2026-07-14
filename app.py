import os
import io
import sys
import uuid
import logging
import tempfile
from flask import Flask, request, jsonify, send_file, send_from_directory
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from invoice_parser import parse_invoice, get_available_models, SUPPORTED_IMAGE_EXTENSIONS
from database import init_db, SessionLocal, InvoiceRecord
from datetime import datetime
from sqlalchemy import or_

# 修复 Windows 终端中文乱码问题
if sys.platform == 'win32':
    try:
        os.system('chcp 65001 >nul 2>&1')
    except Exception:
        pass
    # 确保 stdout/stderr 使用 UTF-8 编码
    if hasattr(sys.stdout, 'reconfigure'):
        sys.stdout.reconfigure(encoding='utf-8')
    if hasattr(sys.stderr, 'reconfigure'):
        sys.stderr.reconfigure(encoding='utf-8')

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')
logger = logging.getLogger(__name__)

app = Flask(__name__, static_folder='static', static_url_path='/invoice')

# 初始化数据库
init_db()

# 上传目录
UPLOAD_FOLDER = tempfile.mkdtemp(prefix='invoice_ocr_')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB


@app.route('/')
def redirect_root():
    from flask import redirect
    return redirect('/invoice/')

@app.route('/invoice/')
def index():
    """返回前端页面"""
    return send_from_directory('static', 'index.html')


# 支持的文件格式
SUPPORTED_EXTENSIONS = {'.pdf'} | SUPPORTED_IMAGE_EXTENSIONS


@app.route('/invoice/api/models', methods=['GET'])
def list_models():
    """返回可用模型列表"""
    return jsonify({'models': get_available_models()})

@app.route('/invoice/api/ledger', methods=['GET'])
def get_ledger():
    """查询台账历史记录"""
    query = request.args.get('query', '').strip()
    db = SessionLocal()
    try:
        if query:
            records = db.query(InvoiceRecord).filter(
                or_(
                    InvoiceRecord.filename.like(f"%{query}%"),
                    InvoiceRecord.invoice_number.like(f"%{query}%"),
                    InvoiceRecord.digital_invoice_number.like(f"%{query}%"),
                    InvoiceRecord.seller_name.like(f"%{query}%"),
                    InvoiceRecord.buyer_name.like(f"%{query}%")
                )
            ).order_by(InvoiceRecord.recognition_time.desc()).all()
        else:
            records = db.query(InvoiceRecord).order_by(InvoiceRecord.recognition_time.desc()).limit(100).all()
        
        result = []
        for r in records:
            result.append({
                "id": r.id,
                "filename": r.filename,
                "invoice_number": r.invoice_number,
                "digital_invoice_number": r.digital_invoice_number,
                "invoice_date": r.invoice_date,
                "seller_name": r.seller_name,
                "buyer_name": r.buyer_name,
                "amount": r.amount,
                "tax_amount": r.tax_amount,
                "total_amount": r.total_amount,
                "recognition_time": r.recognition_time.strftime('%Y-%m-%d %H:%M:%S') if r.recognition_time else ''
            })
        return jsonify({'results': result})
    except Exception as e:
        logger.error(f"台账查询失败: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        db.close()


@app.route('/invoice/api/upload', methods=['POST'])
def upload_invoice():
    """上传并识别发票（支持 PDF 和图片格式）"""
    if 'files' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400

    files = request.files.getlist('files')
    if not files or files[0].filename == '':
        return jsonify({'error': '没有选择文件'}), 400

    # 获取前端选择的模型
    model_name = request.form.get('model', None)

    results = []
    for file in files:
        ext = os.path.splitext(file.filename)[1].lower()
        if ext not in SUPPORTED_EXTENSIONS:
            results.append({
                '文件名': file.filename,
                'error': f'不支持的文件格式: {ext}，仅支持 PDF 和常见图片格式（JPEG、PNG、BMP、WebP、TIFF）'
            })
            continue

        # 保存临时文件（用 UUID 避免并发冲突）
        safe_name = f"{uuid.uuid4().hex}{ext}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], safe_name)
        file.save(filepath)

        try:
            logger.info(f'开始识别: {file.filename} (模型: {model_name or "默认"})')
            result = parse_invoice(filepath, model_name=model_name)
            result['文件名'] = file.filename

            # --- DB Logging & Duplicate Detection ---
            is_duplicate = False
            history = []
            
            db = SessionLocal()
            try:
                # 获取发票号码 (支持全电/数电发票)
                inv_num = result.get('数电发票号码') or result.get('发票号码')
                
                if inv_num:
                    # 查询历史记录
                    existing_records = db.query(InvoiceRecord).filter(
                        or_(
                            InvoiceRecord.invoice_number == inv_num,
                            InvoiceRecord.digital_invoice_number == inv_num
                        )
                    ).all()
                    
                    if existing_records:
                        is_duplicate = True
                        history = [{
                            "filename": rec.filename,
                            "recognition_time": rec.recognition_time.strftime('%Y-%m-%d %H:%M:%S')
                        } for rec in existing_records]
                    
                    # 金额转换辅助方法
                    def parse_float(val):
                        if not val: return 0.0
                        try:
                            # 移除非数字字符
                            return float(str(val).replace(',', '').strip())
                        except ValueError:
                            return 0.0

                    # 插入新记录
                    new_record = InvoiceRecord(
                        filename=file.filename,
                        invoice_number=result.get('发票号码', ''),
                        invoice_date=result.get('发票日期', ''),
                        invoice_type=result.get('发票类型', ''),
                        digital_invoice_number=result.get('数电发票号码', ''),
                        seller_name=result.get('供应商名称', ''),
                        buyer_name=result.get('购买方名称', ''),
                        buyer_tax_id=result.get('购买方纳税人识别号', ''),
                        amount=parse_float(result.get('金额')),
                        tax_amount=parse_float(result.get('税额')),
                        valid_tax_amount=parse_float(result.get('有效抵扣税额')),
                        total_amount=parse_float(result.get('价税合计')),
                    )
                    db.add(new_record)
                    db.commit()
            except Exception as dbe:
                logger.error(f'数据库操作失败: {str(dbe)}')
                db.rollback()
            finally:
                db.close()
            
            result['is_duplicate'] = is_duplicate
            result['history'] = history
            # ----------------------------------------

            results.append(result)
        except Exception as e:
            results.append({
                '文件名': file.filename,
                'error': f'识别失败: {str(e)}'
            })
        finally:
            # 清理临时文件
            if os.path.exists(filepath):
                os.remove(filepath)

    return jsonify({'results': results})


@app.route('/invoice/api/download', methods=['POST'])
def download_excel():
    """将识别结果导出为 Excel"""
    data = request.get_json()
    if not data or 'results' not in data:
        return jsonify({'error': '没有数据'}), 400

    results = data['results']

    # 创建 Excel
    wb = Workbook()
    ws = wb.active
    ws.title = '发票识别结果'

    # 表头
    headers = ['状态', '文件名', '发票日期', '发票类型', '发票号码', '数电发票号码',
               '供应商名称', '购买方名称', '购买方纳税人识别号',
               '金额', '税额', '有效抵扣税额', '价税合计']

    # 表头样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2B5797', end_color='2B5797', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='B0B0B0'),
        right=Side(style='thin', color='B0B0B0'),
        top=Side(style='thin', color='B0B0B0'),
        bottom=Side(style='thin', color='B0B0B0')
    )

    # 写入表头
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    data_font = Font(name='微软雅黑', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    amount_alignment = Alignment(horizontal='right', vertical='center')

    for row_idx, result in enumerate(results, 2):
        if 'error' in result and result.get('error'):
            ws.cell(row=row_idx, column=1, value=result.get('文件名', ''))
            ws.cell(row=row_idx, column=2, value=f"错误: {result['error']}")
            continue

        for col_idx, header in enumerate(headers, 1):
            if header == '状态':
                is_duplicate = result.get('is_duplicate', False)
                value = '疑似重复' if is_duplicate else '正常'
            else:
                value = result.get(header, '')
                
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.border = thin_border

            # 金额列右对齐
            if header in ['金额', '税额', '有效抵扣税额', '价税合计']:
                cell.alignment = amount_alignment
                # 尝试转数值
                try:
                    cell.value = float(value) if value else ''
                    cell.number_format = '#,##0.00'
                except (ValueError, TypeError):
                    pass
            else:
                cell.alignment = data_alignment

    # 设置列宽
    col_widths = [14, 12, 24, 24, 35, 14, 14, 14, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='发票识别结果.xlsx'
    )


if __name__ == '__main__':
    print("=" * 50)
    print("  发票识别系统已启动")
    print("  访问 http://localhost:8999")
    print("=" * 50)
    app.run(host='0.0.0.0', port=8999, debug=True)
